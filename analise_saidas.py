from openpyxl import load_workbook
import analise


pasta = "C:/Users/Nob/Documents/Workspace/AnaliseFinanceira/"


file = pasta + "Nob.xlsx"

# Load spreadsheet
wb = load_workbook(file)

# Load a sheet into a DataFrame by name: df1
sheet = wb['Operacoes']

for i in range(3, sheet.max_row+1):
    #print(i, sheet.cell(row=i, column=1).value)
    ticker = sheet.cell(row=i, column=5).value
    dtinicio = sheet.cell(row=i, column=1).value
    qtd = sheet.cell(row=i, column=8).value
    vCompra = sheet.cell(row=i, column=9).value

    dtinicio = dtinicio.strftime("%Y-%m-%d") 
    #print ("\nAnalisando "+ ticker+' - qtd:'+ str(qtd)+ ' - Valor:'+ str(vCompra))
    analise.AnalisaSaida(ticker, qtd, vCompra, pasta)